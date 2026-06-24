import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from database import get_connection
import business_logic

TEMPLATE_PATH = 'TDG_Masterfile BQLDA_v1_20260623.xlsx'

def get_excel_report_stream():
    # 1. Load Workbook
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
    
    # 2. Get data from Database
    conn = get_connection()
    cursor = conn.cursor()
    
    # Get Master rows
    cursor.execute("SELECT * FROM master_bang_tonghop ORDER BY id ASC")
    master_rows = [dict(row) for row in cursor.fetchall()]
    
    # Get Sub-table rows
    cursor.execute("SELECT * FROM hso_tienkc ORDER BY id ASC")
    hso_rows = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT * FROM kh_thang_tuan ORDER BY id ASC")
    kh_rows = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT * FROM phat_sinh ORDER BY id ASC")
    ps_rows = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT * FROM cu_dac_thu ORDER BY id ASC")
    cu_rows = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT * FROM bu_tien_do ORDER BY id ASC")
    bu_rows = [dict(row) for row in cursor.fetchall()]
    
    conn.close()

    # --- Write to BANG TONG HOP ---
    ws_master = wb['BANG TONG HOP']
    
    # Clear existing data rows below Row 5
    max_row = ws_master.max_row
    if max_row >= 6:
        ws_master.delete_rows(6, max_row - 5)
    
    # Header styling for Warning Column (Col BE / Column 57)
    ws_master.cell(row=3, column=57, value="C. CẢNH BÁO")
    ws_master.cell(row=4, column=57, value="Cảnh báo")
    ws_master.cell(row=3, column=57).font = Font(name="Arial", size=10, bold=True)
    ws_master.cell(row=4, column=57).font = Font(name="Arial", size=9, bold=True)
    
    # Fills
    gray_fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    
    # Fonts
    red_font = Font(name="Arial", size=9, color="FF9C0006", bold=True)
    orange_font = Font(name="Arial", size=9, color="FF9C6500", bold=True)
    yellow_font = Font(name="Arial", size=9, color="FFB2A100", bold=True)
    green_font = Font(name="Arial", size=9, color="FF006100", bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='FFD3D3D3'),
        right=Side(style='thin', color='FFD3D3D3'),
        top=Side(style='thin', color='FFD3D3D3'),
        bottom=Side(style='thin', color='FFD3D3D3')
    )

    for idx, p in enumerate(master_rows):
        row = idx + 6
        metrics = business_logic.calculate_project_metrics(p)
        
        # Write values
        ws_master.cell(row=row, column=1, value=p['TT'])
        ws_master.cell(row=row, column=2, value=p['Ma_BSC'])
        ws_master.cell(row=row, column=3, value=p['Goi_thau'])
        ws_master.cell(row=row, column=4, value=p['Nhom_CT'])
        ws_master.cell(row=row, column=5, value=p['Hang_muc'])
        ws_master.cell(row=row, column=6, value=p['Phu_trach'])
        
        # Format dates nicely
        def parse_dt(dt_str):
            if not dt_str:
                return None
            try:
                return datetime.datetime.strptime(dt_str, '%Y-%m-%d').date()
            except Exception:
                return dt_str

        ws_master.cell(row=row, column=7, value=parse_dt(p['Ngay_BD_YC']))
        ws_master.cell(row=row, column=8, value=parse_dt(p['Ngay_KT_YC']))
        ws_master.cell(row=row, column=9, value=p['Ngan_sach'])
        ws_master.cell(row=row, column=10, value=parse_dt(p['KH_phat_hanh_HSTKTC']))
        ws_master.cell(row=row, column=11, value=p['TT_HSTKTC'])
        ws_master.cell(row=row, column=12, value=p['TT_SPECS'])
        ws_master.cell(row=row, column=13, value=p['TT_BOQ'])
        ws_master.cell(row=row, column=14, value=parse_dt(p['KH_LCNT']))
        ws_master.cell(row=row, column=15, value=p['TT_LCNT'])
        ws_master.cell(row=row, column=16, value=parse_dt(p['KH_Ky_HDCU']))
        ws_master.cell(row=row, column=17, value=p['TT_Ky_HDCU'])
        ws_master.cell(row=row, column=18, value=parse_dt(p['KH_PD_KHCU']))
        ws_master.cell(row=row, column=19, value=p['TT_KHCU'])
        ws_master.cell(row=row, column=20, value=p['Gia_tri_HDCU'])
        
        # U (Col 21) = % HĐCU/NS
        ws_master.cell(row=row, column=21, value=f'=IF(OR(I{row}="",T{row}=""),"",T{row}/I{row})')
        
        ws_master.cell(row=row, column=22, value=parse_dt(p['KH_ky_PLHD']))
        ws_master.cell(row=row, column=23, value=p['TT_Ky_PLHD'])
        ws_master.cell(row=row, column=24, value=parse_dt(p['KH_PD_KHTK']))
        ws_master.cell(row=row, column=25, value=p['TT_KHTK'])
        
        # Z, AA, AB, AC are Gatekeeper Formulas
        if p['Ma_BSC']:
            ws_master.cell(row=row, column=26, value=f'=IF($B{row}="","",IF(AND(OR(K{row}="Đã phát hành",K{row}="Hoàn thiện"),M{row}="Đã bàn giao"),"✔","✘"))')
            ws_master.cell(row=row, column=27, value=f'=IF($B{row}="","",IF(Q{row}="Đã CU","✔","✘"))')
            ws_master.cell(row=row, column=28, value=f'=IF($B{row}="","",IF(Y{row}="Đã duyệt","✔","✘"))')
            ws_master.cell(row=row, column=29, value=f'=IF($B{row}="","",IF(AND(Z{row}="✔",AA{row}="✔",AB{row}="✔"),"ĐỦ ĐK KHỞI CÔNG","THIẾU ĐK"))')
        else:
            ws_master.cell(row=row, column=26, value=None)
            ws_master.cell(row=row, column=27, value=None)
            ws_master.cell(row=row, column=28, value=None)
            ws_master.cell(row=row, column=29, value=None)
            
        ws_master.cell(row=row, column=30, value=parse_dt(p['Ngay_BD_Khoi_Cong']))
        
        # AE (Col 31) = HS tiền KC (duyệt) count
        ws_master.cell(row=row, column=31, value=f'=IF($B{row}="","",COUNTIFS(\'01_HSo TienKC\'!$B:$B,$B{row},\'01_HSo TienKC\'!$J:$J,"Đã duyệt"))')
        
        # Financials from engine (or we can write formulas if we want, but writing engine calculated values works too)
        ws_master.cell(row=row, column=32, value=metrics['Luy_ke_HDCU'])
        ws_master.cell(row=row, column=33, value=metrics['Luy_ke_Phat_sinh'])
        ws_master.cell(row=row, column=34, value=metrics['Total_Cost'])
        
        # AI (Col 35) = Tài liệu KH tháng
        ws_master.cell(row=row, column=35, value=f'=IF($B{row}="","",COUNTIFS(\'02_KH Thang_Tuan\'!$B:$B,$B{row},\'02_KH Thang_Tuan\'!$J:$J,"Đã duyệt")&"/"&COUNTIFS(\'02_KH Thang_Tuan\'!$B:$B,$B{row}))')
        
        # AJ, AK, AL = Pending and active counts
        ws_master.cell(row=row, column=36, value=f'=IF($B{row}="","",COUNTIFS(\'03_Phat sinh\'!$C:$C,$B{row},\'03_Phat sinh\'!$M:$M,"Chờ duyệt"))')
        ws_master.cell(row=row, column=37, value=f'=IF($B{row}="","",COUNTIFS(\'04_CU dac thu\'!$C:$C,$B{row},\'04_CU dac thu\'!$N:$N,"Chờ duyệt"))')
        ws_master.cell(row=row, column=38, value=f'=IF($B{row}="","",COUNTIFS(\'05_Bu tien do\'!$B:$B,$B{row},\'05_Bu tien do\'!$N:$N,"Đang thực hiện"))')
        
        # QA/QC/QS
        ws_master.cell(row=row, column=39, value=p['QA_KH_Thang'])
        ws_master.cell(row=row, column=40, value=p['QA_KQ_Thang'])
        ws_master.cell(row=row, column=41, value=p['QA_Danh_gia_Thang'])
        
        # Construction Progress
        ws_master.cell(row=row, column=42, value=p['KH_Thang'])
        ws_master.cell(row=row, column=43, value=p['KQ_Thang'])
        ws_master.cell(row=row, column=44, value=p['Danh_gia_Thang'])
        
        # Weeks 1-4
        ws_master.cell(row=row, column=45, value=p['T1_KH'])
        ws_master.cell(row=row, column=46, value=p['T1_KQ'])
        ws_master.cell(row=row, column=47, value=p['T1_Danh_gia'])
        
        ws_master.cell(row=row, column=48, value=p['T2_KH'])
        ws_master.cell(row=row, column=49, value=p['T2_KQ'])
        ws_master.cell(row=row, column=50, value=p['T2_Danh_gia'])
        
        ws_master.cell(row=row, column=51, value=p['T3_KH'])
        ws_master.cell(row=row, column=52, value=p['T3_KQ'])
        ws_master.cell(row=row, column=53, value=p['T3_Danh_gia'])
        
        ws_master.cell(row=row, column=54, value=p['T4_KH'])
        ws_master.cell(row=row, column=55, value=p['T4_KQ'])
        ws_master.cell(row=row, column=56, value=p['T4_Danh_gia'])
        
        # Write Warning text into BE (Column 57)
        warning_val = metrics['Co_Canh_bao']
        if p['Ma_BSC']:
            if warning_val == 'RED':
                txt = "ĐỎ"
            elif warning_val == 'ORANGE':
                txt = "CAM"
            elif warning_val == 'YELLOW':
                txt = "VÀNG"
            else:
                txt = "XANH"
            c_warn = ws_master.cell(row=row, column=57, value=txt)
            
            # Apply color-coding directly
            if warning_val == 'RED':
                c_warn.fill = red_fill
                c_warn.font = red_font
            elif warning_val == 'ORANGE':
                c_warn.fill = orange_fill
                c_warn.font = orange_font
            elif warning_val == 'YELLOW':
                c_warn.fill = yellow_fill
                c_warn.font = yellow_font
            else:
                c_warn.fill = green_fill
                c_warn.font = green_font
        else:
            ws_master.cell(row=row, column=57, value=None)

        # Style Read-only automated cells
        read_only_cols = [21, 26, 27, 28, 29, 31, 32, 33, 34, 35, 36, 37, 38]
        for col_idx in read_only_cols:
            ws_master.cell(row=row, column=col_idx).fill = gray_fill
            
        # Borders and Alignments for row
        for c in range(1, 58):
            cell = ws_master.cell(row=row, column=c)
            cell.border = thin_border
            if c in [1, 2, 3, 11, 12, 13, 15, 17, 19, 23, 25, 26, 27, 28, 31, 35, 57]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in [7, 8, 10, 14, 16, 18, 22, 24, 30]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = 'yyyy-mm-dd'
            elif c in [9, 20, 32, 33, 34, 39, 40, 42, 43, 45, 46, 48, 49, 51, 52, 54, 55]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if c not in [39, 40, 42, 43, 45, 46, 48, 49, 51, 52, 54, 55]:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.0%'

    # --- Write Sub-sheets ---
    sub_sheets = {
        '01_HSo TienKC': (hso_rows, [
            ('id', None), ('Ma_BSC', None), ('Hang_muc', None), ('Loai_ho_so', None),
            ('Ten_san_pham', None), ('Link_luu_tru', None), ('Ngay_HT', 'date'),
            ('Nguoi_lap', None), ('Nguoi_duyet', None), ('TT_duyet', None)
        ]),
        '02_KH Thang_Tuan': (kh_rows, [
            ('id', None), ('Ma_BSC', None), ('Hang_muc', None), ('Thang', None),
            ('Loai_tai_lieu', None), ('Noi_dung_chinh', None), ('Dat_YCKT_CDT', None),
            ('Link_tai_lieu', None), ('TT_lap', None), ('TT_duyet', None),
            ('Nguoi_lap', None), ('Nguoi_duyet', None), ('Ngay_duyet', 'date')
        ]),
        '03_Phat sinh': (ps_rows, [
            ('id', None), ('Ma_PS', None), ('Ma_BSC', None), ('Hang_muc', None),
            ('Ngay_PS', 'date'), ('Loai', None), ('Mo_ta', None), ('Nguyen_nhan', None),
            ('De_xuat_xu_ly', None), ('Gia_tri_phat_sinh', 'float'), ('Anh_huong_TD', 'float'),
            ('Link_ho_so', None), ('TT_Phe_duyet', None), ('Nguoi_duyet', None),
            ('Ngay_duyet', 'date'), ('Noi_dung_dieu_chinh', None), ('Ghi_chu', None)
        ]),
        '04_CU dac thu': (cu_rows, [
            ('id', None), ('Ma_YC', None), ('Ma_BSC', None), ('Hang_muc', None),
            ('Ngay_YC', 'date'), ('Loai_YC', None), ('Vat_tu_thiet_bi', None),
            ('Noi_dung_yeu_cau', None), ('KL', 'float'), ('DVT', None),
            ('Gia_tri_phat_sinh', 'float'), ('Trong_Ngoai_HDCU', None), ('Link_ho_so', None),
            ('TT_Phe_duyet', None), ('Nguoi_duyet', None), ('Ngay_can', 'date'),
            ('TT_cung_ung', None), ('Ghi_chu', None)
        ]),
        '05_Bu tien do': (bu_rows, [
            ('id', None), ('Ma_BSC', None), ('Hang_muc', None), ('Ngay_phat_hien', 'date'),
            ('Muc_cham_ngay', 'float'), ('Nguyen_nhan', None), ('Phuong_an', None),
            ('Chi_tiet_giai_phap', None), ('Moc_cam_ket_HT', 'date'), ('Link_phuong_an', None),
            ('TT_duyet', None), ('Nguoi_duyet', None), ('KQ_thuc_hien_bu', None),
            ('TT_Trien_khai', None), ('Ghi_chu', None)
        ])
    }

    import datetime
    for sheet_name, (rows, col_config) in sub_sheets.items():
        ws = wb[sheet_name]
        
        # Clear existing data rows below Row 2
        m_row = ws.max_row
        if m_row >= 3:
            ws.delete_rows(3, m_row - 2)
            
        for r_idx, r_data in enumerate(rows):
            row = r_idx + 3
            
            # STT
            ws.cell(row=row, column=1, value=r_idx + 1)
            
            # Data columns
            for c_idx, (col_name, col_type) in enumerate(col_config[1:]):
                col = c_idx + 2
                val = r_data[col_name]
                
                if col_type == 'date' and val:
                    try:
                        val = datetime.datetime.strptime(val, '%Y-%m-%d').date()
                        cell = ws.cell(row=row, column=col, value=val)
                        cell.number_format = 'yyyy-mm-dd'
                    except Exception:
                        ws.cell(row=row, column=col, value=val)
                elif col_type == 'float' and val is not None:
                    cell = ws.cell(row=row, column=col, value=float(val))
                    cell.number_format = '#,##0.00'
                else:
                    ws.cell(row=row, column=col, value=val)
                    
            # Borders and formats
            for col in range(1, len(col_config) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

    # --- Fix DASHBOARD sheet formulas ---
    ws_dash = wb['DASHBOARD']
    # Repair warning flag COUNTIF formulas in column C
    # C19 = Red warning count
    ws_dash['C19'] = '=COUNTIF(\'BANG TONG HOP\'!BE6:BE100,"ĐỎ")'
    # C20 = Orange warning count
    ws_dash['C20'] = '=COUNTIF(\'BANG TONG HOP\'!BE6:BE100,"CAM")'
    # C21 = Yellow warning count
    ws_dash['C21'] = '=COUNTIF(\'BANG TONG HOP\'!BE6:BE100,"VÀNG")'
    # C22 = Green warning count
    ws_dash['C22'] = '=COUNTIF(\'BANG TONG HOP\'!BE6:BE100,"XANH")'

    # Save to BytesIO stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
